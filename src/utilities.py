"""-------------------------------------------------------
    A MODULE OF SMALL PROCEDURES
-------------------------------------------------------"""

import sys
from datetime import date
import openpyxl as op
from dateutil.relativedelta import relativedelta as rd
import constants as cn


def create_bus_years():
    filepath = cn.FPEXCEL
    wkbbd = op.load_workbook(filepath)
    wksbd = cn.WKSBD
    bdcol = cn.WKSBD_ED_COL_1
    bdrw = cn.WKSBD_ED_ROW_1

    # Create Years

    # Set Year 1 Date
    year_1_dt = wkbbd[wksbd].cell(bdrw, bdcol).value

    # Check if date
    if isinstance(year_1_dt, date):
        pass
    else:
        func_name = create_bus_years.__name__
        mod_name = create_bus_years.__module__
        print(f"Error within: {mod_name} : {func_name} \n"
              "year_1_dt is not a date! \n"
              "Please check, fix and run again!")
        sys.exit()

    # Create Yr2
    year_2 = 2
    year_2_dt = year_1_dt + rd(years=1)
    # Create Yr3
    year_3 = 3
    year_3_dt = year_2_dt + rd(years=1)

    print("Creating Business Years")
    # Update Years
    # Year 2
    wkbbd[wksbd].cell(bdrw + 1, bdcol - 1).value = year_2
    wkbbd[wksbd].cell(bdrw + 1, bdcol).value = year_2_dt
    # Year 3
    wkbbd[wksbd].cell(bdrw + 2, bdcol - 1).value = year_3
    wkbbd[wksbd].cell(bdrw + 2, bdcol).value = year_3_dt

    # Save workbook
    wkbbd.save(filepath)

    print("Business Years Complete!")


if __name__ == "__main__":
    print("This module is intended for import only")
