"""-------------------------------------------------------
    CREATION OF THE CALENDAR
        -- Clears existing data
        -- Creates and inserts new calendar data
-------------------------------------------------------"""

import warnings as wn
import sys
from dateutil.relativedelta import relativedelta as rd
import pandas as pd
import numpy as np
from openpyxl import load_workbook as lw
import constants as cn
from global_functions import (
    day_type as dt,
    get_end_of_month as eom,
    business_year as bus_yr,
    business_month as bus_mon)


def clear_data_wkscldr():
    # --Initialisation--
    filepath = cn.FPEXCEL
    worksheet = cn.WKSCLDR
    colstart = cn.WKSCLDR_ST_COL_1
    colend = cn.WKSCLDR_ED_COL_2
    df_cal = pd.read_excel(filepath, worksheet,
                           usecols=range(colstart, colend))

    print('Clearing Calendar Data')

    # --Procedure--
    # Loop to re-initialise calender st/ed variables
    for aloop in range(1, 3):

        if aloop == 1:
            colst = cn.WKSCLDR_ST_COL_1
            coled = cn.WKSCLDR_ED_COL_1
        else:
            colst = cn.WKSCLDR_ST_COL_2
            coled = cn.WKSCLDR_ED_COL_2

        # set columns
        cols = range(colst, coled)

        # Ignore the data type warnings
        with wn.catch_warnings():
            wn.simplefilter("ignore", category=FutureWarning)

            # Update excel to missing data (NULLS)
            df_cal.iloc[:, cols] = np.nan

            # Select part of df to export
            df_export = df_cal.iloc[:, cols]

            # Clear create data by setting to NULL
            with pd.ExcelWriter(
                    filepath,
                    engine='openpyxl',
                    mode='a',
                    if_sheet_exists='overlay') as writer:

                # Export data and convert to NULLs
                df_export.to_excel(writer, sheet_name=worksheet,
                                   index=False, startcol=colst, startrow=1,
                                   na_rep='', header=False)

    print("Clearing Calendar Data Complete!")


def insert_data_wkscldr():
    # --Initialisation--
    filepath = cn.FPEXCEL
    worksheet = cn.WKSCLDR

    # --Initalise BusData variables--

    # Load workbook
    workbook = lw(filepath, data_only=True)
    wksbd = cn.WKSBD
    bdcol = cn.WKSBD_ED_COL_1
    bdrw = cn.WKSBD_ED_ROW_1
    try:
        start_date = workbook[wksbd].cell(bdrw, bdcol).value
        end_date = eom(workbook[wksbd].cell(
            bdrw + 2, bdcol).value + rd(months=11))
        date_range = pd.date_range(start=start_date, end=end_date, freq='D')
    except ValueError:
        print("Problem with date value. Check workbook to see it has a correct "
              "business year one date")
        sys.exit(1)
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)

    # --Procedure--
    print('Creating Calendar Data')

    # Create calendar dates
    df_newcal = pd.DataFrame(date_range, columns=['Date'])
    # Create calendar year
    df_newcal['Year'] = df_newcal['Date'].dt.year
    # Create calendar month
    df_newcal['Month'] = df_newcal['Date'].dt.month
    # Create calendar Day
    df_newcal['Day'] = df_newcal['Date'].dt.strftime('%a')
    # Create calendar Day type
    df_newcal['DayType'] = df_newcal['Day'].apply(dt)
    # Create calendar Bus Year
    df_newcal['BusYear'] = df_newcal['Date'].apply(bus_yr, workbook=workbook)
    # Create calendar Bus Month
    df_newcal['BusMonth'] = df_newcal['Date'].apply(bus_mon, workbook=workbook)
    # Create short date
    df_newcal['Date'] = pd.to_datetime(
        df_newcal['Date'], format='%d/%m/%Y').dt.date

    # Loop to re-initialise calender st/ed variables
    for aloop in range(1, 3):
        print(f'Exporting Calendar Data {aloop}')
        if aloop == 1:
            colst = cn.WKSCLDR_ST_COL_1
            coled = cn.WKSCLDR_ED_COL_1
            xlout_col = colst
        else:
            colst = cn.WKSCLDR_ED_COL_1
            coled = df_newcal.shape[1]
            xlout_col = cn.WKSCLDR_ST_COL_2

        cols = range(colst, coled)

        df_export = df_newcal.iloc[:, cols]

        # Add data to spreadsheet
        with pd.ExcelWriter(
                filepath,
                engine='openpyxl',
                mode='a',
                if_sheet_exists='overlay') as writer:

            # Export data
            df_export.to_excel(writer, sheet_name=worksheet,
                               index=False, startcol=xlout_col, startrow=1,
                               na_rep='', header=False)

        print(f'Exporting Calendar Data {aloop} Complete!')

    return df_newcal
